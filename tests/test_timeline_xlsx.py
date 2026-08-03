from __future__ import annotations

from pathlib import Path
import json
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from work_agent_core.cli import build_default_tools
from work_agent_core.config import ModelProfile
from work_agent_core.llm import OpenAICompatibleClient
from work_agent_core.timeline_xlsx import (
    apply_timeline_changes,
    inspect_timeline_workbook,
)


WORKSPACE = Path(__file__).resolve().parents[1]
PROFILE = ModelProfile(
    name="timeline-test",
    provider="openai-compatible",
    base_url="https://example.invalid/v1",
    model="test-model",
    api_key_env="UNUSED",
)


def build_sample_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "推进计划"
    sheet["A1"] = "某项目推进安排"
    headers = [
        "节点ID",
        "工作线",
        "时间节点",
        "关键节点 / 验收口径",
        "当前状态",
        "当前进展",
        "下一步动作",
        "责任方",
        "相关材料",
        "最后更新",
        "剩余天数",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor="DCEAF7")
    rows = [
        ["M-001", "可研", "2026-07-31", "形成可研初稿", "推进中", "已完成框架", "补齐测算", "甲方", "可研", "2026-07-29"],
        ["M-002", "上会", "2026-08-07", "形成上会材料", "未开始", "", "等待初稿", "甲方", "上会材料", "2026-07-29"],
    ]
    for row_index, values in enumerate(rows, start=4):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)
        sheet.cell(row=row_index, column=11, value=f'=IF(C{row_index}="","",1)')
    sheet["E4"].fill = PatternFill("solid", fgColor="FFF2CC")
    sheet.row_dimensions[4].height = 42
    sheet.row_dimensions[5].height = 42
    sheet.merge_cells("A7:K7")
    sheet["A7"] = "说明：本行不是时间节点。"
    workbook.save(path)


class TimelineXlsxTests(unittest.TestCase):
    def test_detects_common_timeline_headers_and_lists_normalized_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "timeline.xlsx"
            build_sample_workbook(path)

            result = inspect_timeline_workbook(path)

        self.assertTrue(result["ok"])
        self.assertEqual(result["layout"]["sheet_name"], "推进计划")
        self.assertEqual(result["layout"]["header_row"], 3)
        self.assertEqual(result["layout"]["field_columns"]["planned_date"], 3)
        self.assertEqual(result["layout"]["field_columns"]["title"], 4)
        self.assertEqual(result["rows"][0]["node_id"], "M-001")

    def test_dry_run_previews_without_touching_source(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "timeline.xlsx"
            build_sample_workbook(path)
            before = path.read_bytes()

            result = apply_timeline_changes(
                path,
                changes=[
                    {
                        "action": "update",
                        "match": {"node_id": "M-001"},
                        "values": {"status": "已完成"},
                    }
                ],
                dry_run=True,
            )

            self.assertTrue(result["dry_run"])
            self.assertEqual(result["changes"][0]["before"]["status"], "推进中")
            self.assertEqual(result["changes"][0]["after"]["status"], "已完成")
            self.assertEqual(path.read_bytes(), before)

    def test_batch_update_add_and_soft_delete_preserves_structure(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "timeline.xlsx"
            output = Path(directory) / "timeline-updated.xlsx"
            build_sample_workbook(source)

            result = apply_timeline_changes(
                source,
                output_path=output,
                dry_run=False,
                changes=[
                    {
                        "action": "update",
                        "match": {"node_id": "M-001"},
                        "values": {
                            "status": "已完成",
                            "actual_date": "2026-07-30",
                            "progress": "可研初稿已完成",
                        },
                    },
                    {
                        "action": "add",
                        "values": {
                            "workstream": "设计装修",
                            "planned_date": "2026-08-20",
                            "title": "装修方案定稿",
                            "status": "未开始",
                            "owner": "合作方",
                        },
                    },
                    {
                        "action": "delete",
                        "match": {"node_id": "M-002"},
                    },
                    {
                        "action": "add",
                        "values": {
                            "workstream": "课程建设",
                            "planned_date": "2026-08-25",
                            "title": "完成首批样课评审",
                            "status": "未开始",
                        },
                    },
                ],
                create_missing_columns=True,
            )

            self.assertFalse(result["dry_run"])
            self.assertTrue(output.is_file())
            workbook = load_workbook(output, data_only=False)
            sheet = workbook["推进计划"]
            self.assertEqual(sheet["E4"].value, "已完成")
            self.assertEqual(sheet["F4"].value, "可研初稿已完成")
            self.assertEqual(sheet["E5"].value, "已取消")
            self.assertEqual(sheet["A6"].value, "M-003")
            self.assertEqual(sheet["B6"].value, "设计装修")
            self.assertEqual(sheet["K6"].value, '=IF(C6="","",1)')
            self.assertEqual(sheet["A7"].value, "M-004")
            self.assertEqual(sheet["A8"].value, "说明：本行不是时间节点。")
            self.assertEqual(sheet.row_dimensions[6].height, 42)
            self.assertEqual(sheet["E6"].fill.fgColor.rgb, sheet["E5"].fill.fgColor.rgb)
            self.assertIn("实际完成日期", [sheet.cell(3, column).value for column in range(1, sheet.max_column + 1)])
            self.assertIn("_Friday变更记录", workbook.sheetnames)
            self.assertEqual(workbook["_Friday变更记录"].sheet_state, "hidden")
            workbook.close()

    def test_rejects_non_timeline_workbook_without_explicit_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "not-timeline.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["姓名", "电话", "单位"])
            sheet.append(["张三", "123", "示例公司"])
            workbook.save(path)

            with self.assertRaisesRegex(ValueError, "未能可靠识别时间节点表"):
                inspect_timeline_workbook(path)

    def test_xlsx_skill_exposes_timeline_manager_via_gateway(self) -> None:
        bus = build_default_tools(WORKSPACE, OpenAICompatibleClient(), PROFILE)
        opened = json.loads(
            bus.get_model_tool("sys_skill").handler(
                {"op": "open", "skill_id": "xlsx", "max_chars": 5000}
            )
        )
        names = {item["name"] for item in opened["available_tools"]}
        self.assertIn("manage_timeline_xlsx", names)
        self.assertIn("manage_project_timeline", names)


if __name__ == "__main__":
    unittest.main()
