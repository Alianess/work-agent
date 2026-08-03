from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
import json
import os
import re
import shutil
import tempfile

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet


CANONICAL_ALIASES: dict[str, tuple[str, ...]] = {
    "node_id": ("节点id", "任务id", "事项id", "编号", "序号", "id"),
    "workstream": ("工作线", "阶段", "项目阶段", "模块", "分类", "类别"),
    "planned_date": (
        "计划日期", "时间节点", "节点时间", "计划完成日期", "计划完成时间",
        "截止日期", "截止时间", "完成时限", "日期",
    ),
    "actual_date": ("实际完成日期", "实际日期", "完成日期", "实际完成时间"),
    "title": (
        "关键节点", "节点内容", "任务名称", "工作事项", "主要任务", "事项",
        "工作内容", "任务", "节点", "里程碑",
    ),
    "completion_criteria": ("验收口径", "完成标准", "验收标准", "交付标准"),
    "status": ("当前状态", "节点状态", "任务状态", "进度状态", "状态"),
    "progress": ("当前进展", "最新进展", "进展情况", "工作进展", "进展"),
    "next_action": ("下一步动作", "下一步计划", "下一步", "后续安排", "后续计划"),
    "owner": ("责任方", "责任人", "负责人", "牵头单位", "责任单位", "主责"),
    "materials": ("相关材料", "成果材料", "交付物", "材料", "成果"),
    "updated_at": ("最后更新", "更新时间", "更新日期", "最新更新时间"),
    "notes": ("备注", "说明", "补充说明"),
    "priority": ("优先级", "重要程度", "等级"),
}

DATE_FIELDS = {"planned_date", "actual_date", "updated_at"}
REQUIRED_DETECTION_FIELDS = {"planned_date", "title"}
CHANGELOG_SHEET = "_Friday变更记录"


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_／/|｜:：()（）\-\[\]【】]+", "", text)


NORMALIZED_ALIASES = {
    field: tuple(_normalize_header(alias) for alias in aliases)
    for field, aliases in CANONICAL_ALIASES.items()
}


@dataclass(frozen=True)
class TimelineLayout:
    sheet_name: str
    header_row: int
    field_columns: dict[str, int]
    headers: dict[int, str]
    score: int

    def payload(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "field_columns": self.field_columns,
            "headers": {str(column): name for column, name in self.headers.items()},
            "score": self.score,
        }


def _canonical_field_for_header(value: Any) -> str | None:
    normalized = _normalize_header(value)
    if not normalized:
        return None
    exact: list[tuple[int, str]] = []
    contains: list[tuple[int, str]] = []
    for field, aliases in NORMALIZED_ALIASES.items():
        for alias in aliases:
            if normalized == alias:
                exact.append((len(alias), field))
            elif len(alias) >= 2 and (alias in normalized or normalized in alias):
                contains.append((len(alias), field))
    candidates = exact or contains
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def _layout_for_row(sheet: Worksheet, row: int) -> TimelineLayout:
    headers: dict[int, str] = {}
    field_columns: dict[str, int] = {}
    max_column = min(max(sheet.max_column or 1, 1), 80)
    for column in range(1, max_column + 1):
        raw = sheet.cell(row=row, column=column).value
        if raw is None or str(raw).strip() == "":
            continue
        header = str(raw).strip()
        headers[column] = header
        field = _canonical_field_for_header(header)
        if field and field not in field_columns:
            field_columns[field] = column
    score = len(field_columns)
    if REQUIRED_DETECTION_FIELDS.issubset(field_columns):
        score += 5
    if "status" in field_columns:
        score += 2
    if "node_id" in field_columns:
        score += 1
    return TimelineLayout(sheet.title, row, field_columns, headers, score)


def detect_timeline_layout(
    workbook,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
    field_mapping: dict[str, Any] | None = None,
) -> TimelineLayout:
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在：{sheet_name}")
        sheets = [workbook[sheet_name]]
    else:
        sheets = [sheet for sheet in workbook.worksheets if sheet.title != CHANGELOG_SHEET]

    candidates: list[TimelineLayout] = []
    for sheet in sheets:
        if header_row:
            candidates.append(_layout_for_row(sheet, int(header_row)))
        else:
            for row in range(1, min(max(sheet.max_row or 1, 1), 40) + 1):
                candidates.append(_layout_for_row(sheet, row))
    candidates.sort(key=lambda item: (item.score, len(item.field_columns)), reverse=True)
    if not candidates:
        raise ValueError("工作簿中没有可检查的工作表。")

    layout = candidates[0]
    if field_mapping:
        layout = _apply_field_mapping(workbook[layout.sheet_name], layout, field_mapping)
    if not REQUIRED_DETECTION_FIELDS.issubset(layout.field_columns):
        available = "、".join(layout.headers.values()) or "无"
        raise ValueError(
            "未能可靠识别时间节点表。至少需要“计划日期/时间节点”和“关键节点/工作事项”两类列；"
            f"当前候选工作表 {layout.sheet_name!r} 的表头为：{available}。"
            "可以显式传 sheet_name、header_row 和 field_mapping。"
        )
    return layout


def _apply_field_mapping(
    sheet: Worksheet,
    layout: TimelineLayout,
    field_mapping: dict[str, Any],
) -> TimelineLayout:
    field_columns = dict(layout.field_columns)
    headers = dict(layout.headers)
    normalized_headers = {
        _normalize_header(value): column for column, value in headers.items()
    }
    for raw_field, raw_target in field_mapping.items():
        field = str(raw_field).strip()
        if field not in CANONICAL_ALIASES:
            raise ValueError(f"未知标准字段：{field}")
        if isinstance(raw_target, int) or str(raw_target).isdigit():
            column = int(raw_target)
        else:
            target = _normalize_header(raw_target)
            column = normalized_headers.get(target, 0)
        if column < 1 or column > max(sheet.max_column or 1, 1):
            raise ValueError(f"字段 {field} 无法映射到列：{raw_target}")
        field_columns[field] = column
        headers.setdefault(column, str(sheet.cell(layout.header_row, column).value or raw_target))
    return TimelineLayout(
        layout.sheet_name,
        layout.header_row,
        field_columns,
        headers,
        layout.score,
    )


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _row_is_empty(sheet: Worksheet, row: int, columns: Iterable[int]) -> bool:
    return not any(sheet.cell(row=row, column=column).value not in (None, "") for column in columns)


def _row_is_timeline_data(sheet: Worksheet, row: int, layout: TimelineLayout) -> bool:
    """Exclude title bands, notes, and footers that happen to occupy mapped columns."""
    for field in ("title", "planned_date"):
        column = layout.field_columns.get(field)
        if column and sheet.cell(row=row, column=column).value not in (None, ""):
            return True
    return False


def _row_has_any_value(sheet: Worksheet, row: int) -> bool:
    return any(
        sheet.cell(row=row, column=column).value not in (None, "")
        for column in range(1, max(sheet.max_column or 1, 1) + 1)
    )


def list_timeline_rows(
    sheet: Worksheet,
    layout: TimelineLayout,
    *,
    limit: int = 200,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    columns = sorted(set(layout.field_columns.values()))
    blank_streak = 0
    max_row = min(max(sheet.max_row or layout.header_row, layout.header_row), layout.header_row + 5000)
    for row in range(layout.header_row + 1, max_row + 1):
        if _row_is_empty(sheet, row, columns) or not _row_is_timeline_data(sheet, row, layout):
            blank_streak += 1
            if blank_streak >= 20:
                break
            continue
        blank_streak = 0
        item: dict[str, Any] = {"row": row}
        for field, column in layout.field_columns.items():
            item[field] = _json_value(sheet.cell(row=row, column=column).value)
        rows.append(item)
        if len(rows) >= max(1, min(int(limit), 2000)):
            break
    return rows


def inspect_timeline_workbook(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
    field_mapping: dict[str, Any] | None = None,
    limit: int = 200,
) -> dict[str, Any]:
    source = Path(path).resolve()
    workbook = _load_workbook(source)
    try:
        layout = detect_timeline_layout(
            workbook,
            sheet_name=sheet_name,
            header_row=header_row,
            field_mapping=field_mapping,
        )
        rows = list_timeline_rows(workbook[layout.sheet_name], layout, limit=limit)
        return {
            "ok": True,
            "path": str(source),
            "sheets": workbook.sheetnames,
            "layout": layout.payload(),
            "row_count": len(rows),
            "rows": rows,
        }
    finally:
        workbook.close()


def _load_workbook(path: Path):
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError("仅支持 .xlsx/.xlsm/.xltx/.xltm 工作簿。")
    return load_workbook(
        path,
        data_only=False,
        keep_vba=path.suffix.lower() in {".xlsm", ".xltm"},
        keep_links=True,
    )


def _coerce_value(field: str, value: Any) -> Any:
    if value is None or value == "":
        return None
    if field == "node_id":
        return str(value)
    if field not in DATE_FIELDS or not isinstance(value, str):
        return value
    text = value.strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return value


def _resolve_value_column(
    sheet: Worksheet,
    layout: TimelineLayout,
    key: str,
    *,
    create_missing_columns: bool,
) -> tuple[int, str]:
    field = str(key).strip()
    if field in layout.field_columns:
        return layout.field_columns[field], field

    normalized = _normalize_header(field)
    for column, header in layout.headers.items():
        if _normalize_header(header) == normalized:
            return column, _canonical_field_for_header(header) or field

    if field not in CANONICAL_ALIASES:
        raise ValueError(f"无法识别要修改的字段或表头：{field}")
    if not create_missing_columns:
        raise ValueError(
            f"工作表缺少字段 {field!r}；如需自动新增对应列，请传 create_missing_columns=true。"
        )
    column = max(sheet.max_column or 0, max(layout.headers or {0: ""})) + 1
    header = CANONICAL_ALIASES[field][0]
    sheet.cell(row=layout.header_row, column=column, value=header)
    if column > 1:
        _copy_cell_style(
            sheet.cell(row=layout.header_row, column=column - 1),
            sheet.cell(row=layout.header_row, column=column),
        )
    layout.field_columns[field] = column
    layout.headers[column] = header
    return column, field


def _copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def _copy_row_template(sheet: Worksheet, source_row: int, target_row: int) -> None:
    if source_row <= 0:
        return
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, max(sheet.max_column or 1, 1) + 1):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        _copy_cell_style(source, target)
        if source.data_type == "f" and isinstance(source.value, str):
            try:
                target.value = Translator(
                    source.value,
                    origin=source.coordinate,
                ).translate_formula(target.coordinate)
            except Exception:
                target.value = source.value
    for validation in sheet.data_validations.dataValidation:
        for column in range(1, max(sheet.max_column or 1, 1) + 1):
            source_coordinate = sheet.cell(source_row, column).coordinate
            if source_coordinate in validation.cells:
                validation.add(sheet.cell(target_row, column))


def _insert_row_preserving_merges(sheet: Worksheet, row: int) -> None:
    shifted_merges: list[str] = []
    unchanged_merges: list[str] = []
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= row:
            shifted_merges.append(
                f"{sheet.cell(merged.min_row + 1, merged.min_col).coordinate}:"
                f"{sheet.cell(merged.max_row + 1, merged.max_col).coordinate}"
            )
        elif merged.max_row >= row:
            shifted_merges.append(
                f"{sheet.cell(merged.min_row, merged.min_col).coordinate}:"
                f"{sheet.cell(merged.max_row + 1, merged.max_col).coordinate}"
            )
        else:
            unchanged_merges.append(str(merged))
        sheet.unmerge_cells(str(merged))
    sheet.insert_rows(row, 1)
    for merged in unchanged_merges + shifted_merges:
        sheet.merge_cells(merged)


def _last_data_row(sheet: Worksheet, layout: TimelineLayout) -> int:
    rows = list_timeline_rows(sheet, layout, limit=5000)
    return max((int(item["row"]) for item in rows), default=layout.header_row)


def _next_node_id(sheet: Worksheet, layout: TimelineLayout) -> str:
    column = layout.field_columns.get("node_id")
    if not column:
        return ""
    prefix = "M"
    maximum = 0
    for row in range(layout.header_row + 1, _last_data_row(sheet, layout) + 1):
        value = str(sheet.cell(row=row, column=column).value or "").strip()
        match = re.fullmatch(r"([A-Za-z\u4e00-\u9fff]+)[-_]?(\d+)", value)
        if not match:
            continue
        prefix = match.group(1)
        maximum = max(maximum, int(match.group(2)))
    return f"{prefix}-{maximum + 1:03d}"


def _match_rows(
    sheet: Worksheet,
    layout: TimelineLayout,
    match: dict[str, Any],
) -> list[int]:
    if "row" in match:
        row = int(match["row"])
        if row <= layout.header_row or row > max(sheet.max_row or 0, layout.header_row):
            return []
        return [row]
    candidates = list_timeline_rows(sheet, layout, limit=5000)
    matched: list[int] = []
    for item in candidates:
        ok = True
        for key, expected in match.items():
            if key not in item:
                raise ValueError(f"match 中的字段不可用：{key}")
            actual = item.get(key)
            if str(_json_value(actual) or "").strip() != str(expected or "").strip():
                ok = False
                break
        if ok:
            matched.append(int(item["row"]))
    return matched


def _row_snapshot(sheet: Worksheet, layout: TimelineLayout, row: int) -> dict[str, Any]:
    result = {"row": row}
    for field, column in layout.field_columns.items():
        result[field] = _json_value(sheet.cell(row=row, column=column).value)
    return result


def _write_values(
    sheet: Worksheet,
    layout: TimelineLayout,
    row: int,
    values: dict[str, Any],
    *,
    create_missing_columns: bool,
) -> None:
    for key, value in values.items():
        column, field = _resolve_value_column(
            sheet,
            layout,
            str(key),
            create_missing_columns=create_missing_columns,
        )
        cell = sheet.cell(row=row, column=column)
        cell.value = _coerce_value(field, value)
        if field in DATE_FIELDS and isinstance(cell.value, (date, datetime)):
            cell.number_format = "yyyy-mm-dd"


def _append_change_log(
    workbook,
    *,
    source_path: Path,
    target_path: Path,
    changes: list[dict[str, Any]],
    change_source: str,
) -> None:
    if CHANGELOG_SHEET in workbook.sheetnames:
        sheet = workbook[CHANGELOG_SHEET]
    else:
        sheet = workbook.create_sheet(CHANGELOG_SHEET)
        sheet.append(["时间", "来源", "源文件", "目标文件", "操作", "工作表", "行", "修改前", "修改后"])
        sheet.sheet_state = "hidden"
    timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    for change in changes:
        sheet.append(
            [
                timestamp,
                change_source,
                str(source_path),
                str(target_path),
                change.get("action"),
                change.get("sheet_name"),
                change.get("row"),
                json.dumps(change.get("before"), ensure_ascii=False, default=str),
                json.dumps(change.get("after"), ensure_ascii=False, default=str),
            ]
        )


def apply_timeline_changes(
    path: str | Path,
    *,
    changes: list[dict[str, Any]],
    output_path: str | Path | None = None,
    sheet_name: str | None = None,
    header_row: int | None = None,
    field_mapping: dict[str, Any] | None = None,
    create_missing_columns: bool = False,
    dry_run: bool = True,
    backup: bool = True,
    record_history: bool = True,
    change_source: str = "Friday",
) -> dict[str, Any]:
    source = Path(path).resolve()
    target = Path(output_path).resolve() if output_path else source
    source_stat = source.stat()
    workbook = _load_workbook(source)
    applied: list[dict[str, Any]] = []
    try:
        layout = detect_timeline_layout(
            workbook,
            sheet_name=sheet_name,
            header_row=header_row,
            field_mapping=field_mapping,
        )
        sheet = workbook[layout.sheet_name]
        for raw_change in changes:
            if not isinstance(raw_change, dict):
                raise ValueError("changes 中的每一项都必须是对象。")
            action = str(raw_change.get("action") or "").strip().lower()
            values = raw_change.get("values") or {}
            if not isinstance(values, dict):
                raise ValueError("change.values 必须是对象。")

            if action == "add":
                row = _last_data_row(sheet, layout) + 1
                if row <= max(sheet.max_row or 0, 0) and _row_has_any_value(sheet, row):
                    _insert_row_preserving_merges(sheet, row)
                    _copy_row_template(sheet, row - 1, row)
                elif row > max(sheet.max_row or 0, 0):
                    _copy_row_template(sheet, row - 1, row)
                elif row > layout.header_row + 1:
                    # Reuse a formatted blank row inside an existing workbook, but
                    # still inherit formulas and validation from the prior node.
                    _copy_row_template(sheet, row - 1, row)
                if "node_id" in layout.field_columns and "node_id" not in values:
                    values = {"node_id": _next_node_id(sheet, layout), **values}
                before = {"row": row}
                _write_values(
                    sheet,
                    layout,
                    row,
                    values,
                    create_missing_columns=create_missing_columns,
                )
                after = _row_snapshot(sheet, layout, row)
            elif action in {"update", "delete"}:
                match = raw_change.get("match") or {}
                if not isinstance(match, dict) or not match:
                    raise ValueError(f"{action} 操作必须提供非空 match。")
                rows = _match_rows(sheet, layout, match)
                if len(rows) != 1:
                    raise ValueError(
                        f"{action} 操作要求唯一命中一行，当前命中 {len(rows)} 行：{rows[:20]}"
                    )
                row = rows[0]
                before = _row_snapshot(sheet, layout, row)
                if action == "update":
                    _write_values(
                        sheet,
                        layout,
                        row,
                        values,
                        create_missing_columns=create_missing_columns,
                    )
                    after = _row_snapshot(sheet, layout, row)
                else:
                    delete_mode = str(raw_change.get("delete_mode") or "soft").strip().lower()
                    if delete_mode == "row":
                        sheet.delete_rows(row, 1)
                        after = {"row": row, "deleted": True}
                    elif delete_mode == "soft":
                        if "status" not in layout.field_columns:
                            raise ValueError("软删除需要状态列；可改用 delete_mode='row'。")
                        soft_values = dict(values)
                        if not soft_values.get("status"):
                            soft_values["status"] = "已取消"
                        _write_values(
                            sheet,
                            layout,
                            row,
                            soft_values,
                            create_missing_columns=create_missing_columns,
                        )
                        after = _row_snapshot(sheet, layout, row)
                    else:
                        raise ValueError("delete_mode 仅支持 soft 或 row。")
            else:
                raise ValueError(f"不支持的时间节点操作：{action or '（空）'}")

            applied.append(
                {
                    "action": action,
                    "sheet_name": layout.sheet_name,
                    "row": row,
                    "before": before,
                    "after": after,
                }
            )

        result = {
            "ok": True,
            "dry_run": bool(dry_run),
            "source_path": str(source),
            "output_path": None if dry_run else str(target),
            "layout": layout.payload(),
            "changes": applied,
        }
        if dry_run:
            return result

        current_stat = source.stat()
        if (
            current_stat.st_mtime_ns != source_stat.st_mtime_ns
            or current_stat.st_size != source_stat.st_size
        ):
            raise RuntimeError("源工作簿在处理期间发生变化，已停止写入以避免覆盖新内容。")

        target.parent.mkdir(parents=True, exist_ok=True)
        if record_history:
            _append_change_log(
                workbook,
                source_path=source,
                target_path=target,
                changes=applied,
                change_source=str(change_source or "Friday"),
            )

        backup_path: Path | None = None
        if target.exists() and backup:
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup_path = target.with_name(f"{target.stem}.backup-{timestamp}{target.suffix}")
            shutil.copy2(target, backup_path)

        with tempfile.NamedTemporaryFile(
            prefix=f".{target.stem}-",
            suffix=target.suffix,
            dir=target.parent,
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
        try:
            workbook.save(temporary)
            os.replace(temporary, target)
        finally:
            temporary.unlink(missing_ok=True)
        result["backup_path"] = str(backup_path) if backup_path else None
        return result
    finally:
        workbook.close()


def manage_timeline_xlsx(args: dict[str, Any]) -> dict[str, Any]:
    operation = str(args.get("operation") or "inspect").strip().lower()
    common = {
        "sheet_name": str(args.get("sheet_name") or "").strip() or None,
        "header_row": int(args["header_row"]) if args.get("header_row") else None,
        "field_mapping": args.get("field_mapping") or None,
    }
    if operation in {"inspect", "list"}:
        return inspect_timeline_workbook(
            args["path"],
            limit=int(args.get("limit") or 200),
            **common,
        )
    if operation == "apply":
        changes = args.get("changes")
        if not isinstance(changes, list) or not changes:
            raise ValueError("apply 操作必须提供非空 changes 数组。")
        return apply_timeline_changes(
            args["path"],
            changes=changes,
            output_path=args.get("output_path"),
            create_missing_columns=bool(args.get("create_missing_columns")),
            dry_run=bool(args.get("dry_run", True)),
            backup=bool(args.get("backup", True)),
            record_history=bool(args.get("record_history", True)),
            change_source=str(args.get("change_source") or "Friday"),
            **common,
        )
    raise ValueError("operation 仅支持 inspect、list 或 apply。")
