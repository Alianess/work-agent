"""Create an .xlsx workbook from Markdown tables.

Each Markdown table becomes a worksheet. The sheet name is taken from the
nearest preceding H2/H3 heading; if none, sheets are named Sheet1, Sheet2, ...
Requires `openpyxl`. Run via the office_python interpreter (see skill_runtime).
"""

from __future__ import annotations

from pathlib import Path
import argparse
import json
import re


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    markdown = _read_markdown(args)
    if not markdown.strip():
        raise ValueError("Markdown 内容为空。")
    output_path = Path(args.output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    create_xlsx(markdown, output_path, sheet_name=args.sheet_name or "")
    print(
        json.dumps(
            {"output_path": str(output_path), "sheets": _count_tables(markdown)},
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create an .xlsx workbook from Markdown tables.")
    parser.add_argument("--markdown-path")
    parser.add_argument("--markdown-content")
    parser.add_argument("--output-path", required=True)
    parser.add_argument("--sheet-name", default="")
    return parser.parse_args(argv)


def _read_markdown(args: argparse.Namespace) -> str:
    if args.markdown_path:
        return Path(args.markdown_path).read_text(encoding="utf-8", errors="replace")
    return args.markdown_content or ""


def create_xlsx(markdown: str, output_path: Path, *, sheet_name: str = "") -> None:
    from openpyxl import Workbook

    tables = _split_tables(markdown)
    if not tables:
        raise ValueError("未在 Markdown 中找到任何表格。")

    workbook = Workbook()
    # Remove the default sheet; we recreate per table.
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for index, (heading, rows) in enumerate(tables, start=1):
        name = (sheet_name if index == 1 else "") or heading or f"Sheet{index}"
        name = _sanitize_sheet_name(name)[:31]
        if name in workbook.sheetnames:
            name = f"{name[:27]}-{index}"
        worksheet = workbook.create_sheet(title=name)
        for row_index, row in enumerate(rows, start=1):
            for col_index, cell_value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=_coerce(cell_value))
                if row_index == 1:
                    cell.font = cell.font.copy(bold=True)
    workbook.save(str(output_path))


def _split_tables(markdown: str) -> list[tuple[str, list[list[str]]]]:
    """Split markdown into (heading, rows) per table. Skips separator rows."""
    lines = markdown.splitlines()
    tables: list[tuple[str, list[list[str]]]] = []
    current_heading = ""
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if stripped.startswith("#"):
            current_heading = stripped.lstrip("#").strip()
            i += 1
            continue
        if stripped.startswith("|") and stripped.endswith("|"):
            # Collect contiguous table rows.
            rows: list[list[str]] = []
            heading = current_heading
            while i < len(lines) and lines[i].strip().startswith("|") and lines[i].strip().endswith("|"):
                row_line = lines[i].strip()
                cells = [cell.strip() for cell in row_line[1:-1].split("|")]
                # Skip separator rows (---|:|---)
                if all(re.fullmatch(r":?-{2,}:?", c) for c in cells):
                    i += 1
                    continue
                rows.append(cells)
                i += 1
            if rows:
                tables.append((heading, rows))
            continue
        i += 1
    return tables


def _count_tables(markdown: str) -> int:
    return len(_split_tables(markdown))


def _coerce(value: str):
    if value == "" or value is None:
        return None
    if re.fullmatch(r"-?\d+", value):
        try:
            return int(value)
        except ValueError:
            pass
    if re.fullmatch(r"-?\d+\.\d+", value):
        try:
            return float(value)
        except ValueError:
            pass
    if value.lower() == "true":
        return True
    if value.lower() == "false":
        return False
    # Strip surrounding quotes
    if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
        return value[1:-1]
    return value


def _sanitize_sheet_name(name: str) -> str:
    # Excel disallows : \ / ? * [ ]
    return re.sub(r"[\[\]:\\/*?]", "-", name).strip() or "Sheet"


if __name__ == "__main__":
    raise SystemExit(main())
